from __future__ import annotations
import numpy as np


def read_excel(path: str, cells: str, sheet_name: str = None) -> np.ndarray:
    """Read an Excel file and return an array of cell values."""

    from openpyxl import load_workbook

    # open excel file
    wb = load_workbook(filename=path)
    ws = wb.active if sheet_name is None else wb[sheet_name]

    # read cells
    cells_tuple = ws[cells]
    get_values = np.vectorize(lambda cell: cell.value, otypes=[object])

    return get_values(cells_tuple)


def visualize(process_time, start_time) -> None:
    """
    Visualize the scheduling of a production line problem.

    Parameters:
        process_time: A table representing the processing time of the jobs on different machines,
            where the rows are the jobs and columns are the machines.
        start_time: A table representing the time to start processing the jobs on different machines,
            where the rows are the jobs and columns are the machines.
    """

    # NOTE: Oops. I think I over-engineered the code... ;-;

    import pygame
    import random
    import time
    from typing import Any

    # input validation
    process_time = np.array(process_time)
    start_time = np.array(start_time)

    if process_time.shape != start_time.shape or np.any(
        (process_time[process_time != None] < 0) | (start_time[start_time != None] < 0)
    ):
        raise ValueError(
            "The two input tables should be of the same dimensions and all values must be non-negative or None."
        )

    # process input data
    ignore_none = np.vectorize(
        lambda op, a, b: None if a is None or b is None else op(a, b), otypes=[object]
    )

    process_time: np.ndarray = ignore_none(round, process_time, 10)
    start_time: np.ndarray = ignore_none(round, start_time, 10)
    end_time: np.ndarray = ignore_none(lambda a, b: a + b, start_time, process_time)

    num_jobs, num_machines = process_time.shape

    # convert to a sorted list
    start_time_list = [
        (t, j, m)
        for j in range(num_jobs)
        for m in range(num_machines)
        if (t := start_time[j][m]) is not None
    ]
    start_time_list.sort()

    end_time_list = [
        (t, j, m)
        for j in range(num_jobs)
        for m in range(num_machines)
        if (t := end_time[j][m]) is not None
    ]
    end_time_list.sort()

    max_time: float
    max_time, _, _ = end_time_list[-1]

    # pygame initialization
    pygame.init()
    pygame.display.set_caption("Scheduling of a production line")

    resolution = (1000, 600)
    screen = pygame.display.set_mode(resolution)

    white = (255, 255, 255)
    font = pygame.font.SysFont("Courier New", 50, bold=True)

    # components definition
    mach_size = 124
    job_size = 110

    class Component:
        def __init__(
            self,
            surface: pygame.Surface,
            pos: tuple[int, int] = None,
            children: list[Component] = None,
        ):
            self._surface = surface
            self._pos = np.array(pos) if pos is not None else np.zeros(2)
            self._global_pos = self._pos

            self._parent: Component = None
            self._children: list[Component] = []
            self.add_children(children)

            self._movement = None

            self.set_dirty()

        def __iter__(self):
            return iter(self._children)

        @property
        def pos(self) -> np.ndarray:
            return self._pos

        @pos.setter
        def pos(self, value):
            if np.array_equal(self._pos, value):
                return
            self._pos = np.array(value)
            self._update_global_pos()
            if self._parent is not None:
                self._parent.set_dirty()

        @property
        def global_pos(self) -> np.ndarray:
            return self._global_pos

        @global_pos.setter
        def global_pos(self, value):
            if np.array_equal(self._global_pos, value):
                return
            self.pos = (
                value - self._parent.global_pos if self._parent is not None else value
            )

        def _update_global_pos(self):
            curr_global_pos = (
                self._parent.global_pos + self.pos
                if self._parent is not None
                else self.pos
            )
            # return if does not change
            if np.array_equal(self.global_pos, curr_global_pos):
                return
            self._global_pos = curr_global_pos
            self._on_global_pos_updated()
            for child in self:
                child._update_global_pos()

        def _on_global_pos_updated(self):
            pass

        def set_dirty(self):
            """Set the dirty flag for this component. Dirty components are repainted every rendering update."""
            self._is_dirty = True

        def render(self, surface: pygame.Surface):
            surface.blit(self._surface, self.pos)

        def repaint(self):
            """Override this method in child classes to implement the components' look.
            There must be a call to super().repaint() at the end of the method."""

            # clear first if it's an empty Component
            if type(self) is Component:
                self._surface.fill((0, 0, 0, 0))

            for child in self:
                child.render(self._surface)

        def render_update(self):
            """Call this method on the root component every frame to process rendering updates."""

            # recursively checking the chain from bottom up
            for child in self:
                child.render_update()

            if self._is_dirty:
                self.repaint()
                self._is_dirty = False
                if self._parent is not None:
                    # setting dirty up the parent to repaint the affected chain
                    self._parent.set_dirty()

        def add_child(self, component: Component):
            if component is None:
                return
            self._children.append(component)
            component._parent = self
            component._update_global_pos()
            self.set_dirty()

        def add_children(self, components: list[Component]):
            if not components:
                return
            self._children.extend(components)
            for component in components:
                component._parent = self
                component._update_global_pos()
            self.set_dirty()

        def update(self):
            """Override this method in child classes to implement components' update logic.
            There must a call to super().update() at the beginning of the method."""

            for child in self:
                child.update()

            if self._movement is not None:
                target, [duration, prev_time] = self._movement

                curr_time = time.time()
                delta_time = curr_time - prev_time
                progress = delta_time / duration

                if progress >= 1:
                    # passed the movement duration
                    self.pos = target
                    self._movement = None
                else:
                    # update pos
                    self.pos = self.pos + (target - self.pos) * progress

                    # update the movement object
                    self._movement[1][0] = duration - delta_time
                    self._movement[1][1] = curr_time

        def move_pos(self, target, duration: float = 0.0):
            target = np.array(target)
            if duration > 0:
                # record a movement object to be processed in update()
                self._movement = target, [duration, time.time()]
            else:
                self.pos = target
                self._movement = None

        def get_size(self) -> np.ndarray:
            return np.array(self._surface.get_size())

        def get_pos_center(
            self, target: pygame.Surface | tuple[int, int]
        ) -> np.ndarray:
            size = self.get_size()
            return np.round(
                (target.get_size() - size) / 2
                if isinstance(target, pygame.Surface)
                else target - size / 2
            )

    class Label(Component):
        def __init__(
            self,
            text: str = "Label",
            color: tuple[int, int, int] = white,
            font: pygame.font.Font = font,
            pos: tuple[int, int] = None,
            children: list[Component] = None,
        ):
            super().__init__(None, pos, children)
            self._text = text
            self._color = color
            self._font = font

            # pre-paint before the rendering update frame to ensure the surface is
            # available after this constructor.
            self.repaint()
            self._is_dirty = False

        def repaint(self):
            self._surface = self._font.render(self._text, True, self._color)
            super().repaint()

        @property
        def text(self) -> str:
            return self._text

        @text.setter
        def text(self, value: str):
            if self._text == value:
                return
            self._text = value
            self.set_dirty()

    class LabeledComponent(Component):
        def __init__(
            self,
            label: str,
            fgcolor: tuple[int, int, int],
            font: pygame.font.Font,
            bgcolor: tuple[int, int, int],
            border_color: tuple[int, int, int],
            size: tuple[int, int],
            pos: tuple[int, int],
            children: list[Component],
            is_background_alpha: bool = False,
        ):
            flags = pygame.SRCALPHA if is_background_alpha else 0
            surface = pygame.Surface(size, flags)

            super().__init__(surface, pos, children)

            lbl = Label(label, fgcolor, font)
            lbl.pos = lbl.get_pos_center(surface)
            self._label = lbl

            self._bgcolor = bgcolor
            self._border_color = border_color

        def repaint(self):
            self._label.render(self._surface)
            super().repaint()

    class Machine(LabeledComponent):
        def __init__(
            self,
            label: str = "M",
            fgcolor: tuple[int, int, int] = white,
            font: pygame.font.Font = font,
            bgcolor: tuple[int, int, int] = (128, 25, 59),
            border_color: tuple[int, int, int] = (80, 0, 47),
            size: tuple[int, int] = (mach_size, mach_size),
            pos: tuple[int, int] = None,
            children: list[Component] = None,
        ):
            super().__init__(
                label, fgcolor, font, bgcolor, border_color, size, pos, children
            )

        def repaint(self):
            self._surface.fill(self._bgcolor)
            if self._border_color is not None:
                pygame.draw.rect(
                    self._surface, self._border_color, self._surface.get_rect(), width=9
                )
            super().repaint()

    class Job(LabeledComponent):
        def __init__(
            self,
            label: str = "J",
            fgcolor: tuple[int, int, int] = white,
            font: pygame.font.Font = font,
            bgcolor: tuple[int, int, int] = (0, 158, 26),
            border_color: tuple[int, int, int] = (0, 75, 12),
            size: tuple[int, int] = (job_size, job_size),
            pos: tuple[int, int] = None,
            children: list[Component] = None,
        ):
            super().__init__(
                label,
                fgcolor,
                font,
                bgcolor,
                border_color,
                size,
                pos,
                children,
                is_background_alpha=True,
            )

        def repaint(self):
            self._surface.fill((0, 0, 0, 0))  # clear the surface
            rect = self._surface.get_rect()
            pygame.draw.ellipse(self._surface, self._bgcolor, rect)
            if self._border_color is not None:
                pygame.draw.ellipse(self._surface, self._border_color, rect, width=6)
            super().repaint()

        @property
        def color(self):
            return self._bgcolor

    class Button(Component):
        NONE = 0
        HOVERING = 1
        PRESSING = 2

        def __init__(
            self,
            size: tuple[int, int] = None,
            pos: tuple[int, int] = None,
            children: list[Component] = None,
        ):
            self._size = size if size is not None else (35, 35)
            surface = pygame.Surface(self._size)
            super().__init__(surface, pos, children)
            self._rect = pygame.Rect(self.global_pos, self._size)
            self._state = Button.NONE

        def repaint(self):
            if self._state == Button.HOVERING:
                self._surface.fill((150, 150, 150))
            elif self._state == Button.PRESSING:
                self._surface.fill((120, 120, 120))
            else:
                self._surface.fill((169, 169, 169))
            super().repaint()

        def _on_global_pos_updated(self):
            self._rect.update(self.global_pos, self._size)

        def update(self):
            super().update()

            mouse_pos = pygame.mouse.get_pos()
            collide = self._rect.collidepoint(mouse_pos)
            pressed = pygame.mouse.get_pressed()

            if self._state != Button.NONE and (
                not collide or not pygame.mouse.get_focused()
            ):
                self._state = Button.NONE
            elif collide and pressed[0] and self._state == Button.HOVERING:
                self._state = Button.PRESSING
            elif collide and not pressed[0] and self._state != Button.HOVERING:
                if self._state == Button.PRESSING:
                    # fire callback if it goes from PRESSING to HOVERING
                    self.on_clicked()
                self._state = Button.HOVERING
            else:
                return

            self.set_dirty()

        def on_clicked(self):
            pass

    class PlayButton(Button):
        def __init__(
            self,
            size: tuple[int, int] = None,
            pos: tuple[int, int] = None,
            children: list[Component] = None,
        ):
            super().__init__(size, pos, children)
            self._play_polygon = ((10, 6), (10, 29), (25, 17))
            self._pause_lines = ((11, 6), (11, 29), (24, 6), (24, 29))
            self._colors = ((0, 120, 0), (120, 0, 0))
            self._is_playing = False

        @property
        def is_playing(self):
            return self._is_playing

        @is_playing.setter
        def is_playing(self, value: bool):
            if self._is_playing == value:
                return
            self._is_playing = value
            self.set_dirty()

        def repaint(self):
            super().repaint()
            if self.is_playing:
                _, color = self._colors
                line1a, line1b, line2a, line2b = self._pause_lines
                pygame.draw.line(self._surface, color, line1a, line1b, 7)
                pygame.draw.line(self._surface, color, line2a, line2b, 7)
            else:
                color, _ = self._colors
                pygame.draw.polygon(self._surface, color, self._play_polygon)

        def on_clicked(self):
            self.is_playing = not self.is_playing

    class ResetButton(Button):
        def __init__(
            self,
            size: tuple[int, int] = None,
            pos: tuple[int, int] = None,
            children: list[Component] = None,
        ):
            super().__init__(size, pos, children)

            reset_symbol = pygame.Surface((35, 35), pygame.SRCALPHA)
            color = (120, 0, 0)
            pygame.draw.ellipse(reset_symbol, color, (7, 7, 21, 21), 4)
            pygame.draw.rect(reset_symbol, (0, 0, 0, 0), (7, 7, 10, 10))
            pygame.draw.polygon(reset_symbol, color, ((17, 3), (17, 14), (10, 8)))

            self._reset_symbol = reset_symbol

        def repaint(self):
            super().repaint()
            self._surface.blit(self._reset_symbol, (0, 0))

    class ProgressBar(Component):
        def __init__(
            self,
            length: int,
            pos: tuple[int, int] = None,
            children: list[Component] = None,
        ):
            surface = pygame.Surface((length, 160), pygame.SRCALPHA)
            super().__init__(surface, pos, children)
            self._length = length
            self._progress = 0.0

        @property
        def progress(self):
            return self._progress

        @progress.setter
        def progress(self, value: float):
            # clamp value between [0, 1]
            value = 0.0 if value < 0 else 1.0 if value > 1 else value
            if self._progress == value:
                return
            self._progress = value
            self.set_dirty()

        def repaint(self):
            self._surface.fill((0, 0, 0, 0))

            line_height = 140
            radius = 7
            point_x = self.progress * (self._length - 2 * radius) + radius

            pygame.draw.line(
                self._surface, (200, 0, 0), (point_x, 0), (point_x, 120), 3
            )

            pygame.draw.line(
                self._surface,
                (180, 0, 0),
                (radius, line_height),
                (point_x, line_height),
                5,
            )
            pygame.draw.line(
                self._surface,
                (140, 140, 140),
                (point_x, line_height),
                (self._length - radius, line_height),
                5,
            )
            pygame.draw.circle(
                self._surface, (200, 0, 0), (point_x, line_height), radius
            )

            super().repaint()

    class Slider(Component):
        def __init__(
            self,
            size: tuple[int, int],
            range: tuple[float, float],
            step: float = 1.0,
            default: float = None,
            pos: tuple[int, int] = None,
            children: list[Component] = None,
        ):
            self._size = size
            surface = pygame.Surface(self._size, pygame.SRCALPHA)
            super().__init__(surface, pos, children)

            self._range = range
            self._step = step
            self._value = None
            self.value = default if default is not None else range[0]

            self._rect = pygame.Rect(self.global_pos, self._size)
            self._is_pressing = False
            self._is_dragging = False

        @property
        def value(self) -> float:
            return self._value

        @value.setter
        def value(self, value: float):
            min, max = self._range
            step = self._step

            if step != 0:
                # ensure max is a multiple of step plus min
                max = int((max - min) / step) * step + min

            value = (
                # clamp value between range
                min
                if value < min
                else max
                if value > max
                # no constraint if no step
                else value
                if step == 0
                # else ensure value is a mutiple of step plus min
                else round((value - min) / step) * step + min
            )

            if self._value == value:
                return
            self._value = value
            self.on_value_changed(value)
            self.set_dirty()

        def repaint(self):
            width, height = self._size
            value_x = np.interp(self.value, self._range, (0, width))
            pygame.draw.rect(self._surface, (200, 200, 200), (0, 0, value_x, height))
            pygame.draw.rect(
                self._surface, (40, 40, 40), (value_x, 0, width - value_x, height)
            )
            super().repaint()

        def _on_global_pos_updated(self):
            self._rect.update(self.global_pos, self._size)

        def update(self):
            super().update()

            mouse_pos = pygame.mouse.get_pos()
            collide = self._rect.collidepoint(mouse_pos)
            pressed = pygame.mouse.get_pressed()

            if collide and pressed[0] and not self._is_pressing:
                self._is_dragging = True
            elif not pressed[0] and self._is_dragging:
                self._is_dragging = False
            self._is_pressing = pressed[0]

            if self._is_dragging:
                relative_pos = mouse_pos - self.global_pos
                value = np.interp(relative_pos[0], (0, self._size[0]), self._range)
                self.value = value

        def on_value_changed(self, value):
            pass

    class SchedulingChart(Component):
        def __init__(
            self,
            size: tuple[int, int],
            pos: tuple[int, int] = None,
            children: list[Component] = None,
        ):
            surface = pygame.Surface(size, pygame.SRCALPHA)
            super().__init__(surface, pos, children)

            # draw the chart
            y = 0
            height = size[1] / num_machines
            for m in range(num_machines):
                for j in range(num_jobs):
                    start = start_time[j][m]
                    proc = process_time[j][m]
                    if start is None or proc is None:
                        continue

                    color = [
                        darkened if (darkened := c - 30) >= 0 else 0
                        for c in jobs[j].color
                    ]
                    x = (start / max_time) * size[0]
                    width = (proc / max_time) * size[0]
                    pygame.draw.rect(surface, color, (x, y, width, height))
                y += height

    # random color
    random.seed(5)

    def get_color():
        return tuple(random.randint(0, 255) for _ in range(3))

    # machine and job components position
    def get_spreaded_1d_pos_array(
        border: tuple[int, int] | int, size: int, num: int, margin: int = None
    ) -> tuple[np.ndarray, Any] | np.ndarray:
        """Helper function for generating the position array."""
        if margin is None:
            a, b = border
            arr, step = np.linspace(a, b - size, num, retstep=True)
            return arr, step - size
        else:
            try:
                a, b = border
                total_length = size * num + margin * (num - 1)
                offset = a + ((b - a) - total_length) / 2
            except TypeError:
                offset = border

            arr = np.arange(num) * (size + margin)
            return arr + offset

    def get_pos_array(
        border, size, y_start, total, num_per_row, margin_y
    ) -> tuple[np.ndarray, Any]:
        num_whole_rows = int(total / num_per_row)
        num_rem = total % num_per_row

        margin = 50  # default margin

        arr = np.empty((0, 2))

        y = y_start
        if num_whole_rows > 0:
            x_arr, margin = get_spreaded_1d_pos_array(border, size, num_per_row)
            x_arr = x_arr.reshape((num_per_row, 1))

            # iterate through the rows and append the coordinates
            for _ in range(num_whole_rows):
                y_arr = np.full(x_arr.shape, y)

                row_coord_arr = np.append(x_arr, y_arr, 1)
                arr = np.append(arr, row_coord_arr, 0)

                y += size + margin_y

        if num_rem > 0:
            x_arr = get_spreaded_1d_pos_array(border, size, num_rem, margin).reshape(
                (num_rem, 1)
            )
            y_arr = np.full(x_arr.shape, y)

            row_coord_arr = np.append(x_arr, y_arr, 1)
            arr = np.append(arr, row_coord_arr, 0)

        return arr, y + size

    width, _ = resolution
    margin_y = 30

    border = (100, width - 100)
    num_per_row = 5

    machines_pos, y_end = get_pos_array(
        border, mach_size, 60, num_machines, num_per_row, margin_y
    )
    jobs_pos, _ = get_pos_array(
        border, job_size, y_end + 40, num_jobs, num_per_row, margin_y
    )
    processing_pos = machines_pos + (mach_size - job_size) / 2

    # root component
    content = Component(screen)

    # create components
    def new_machine(i, pos):
        """Helper function to generate Machine"""
        return Machine(f"M{i + 1}", pos=pos)

    def new_job(i, pos):
        """Helper function to generate Job with random color"""
        bg = get_color()
        border = [darkened if (darkened := c - 50) >= 0 else 0 for c in bg]
        return Job(f"J{i + 1}", bgcolor=bg, border_color=border, pos=pos)

    machines = [new_machine(i, machines_pos[i]) for i in range(num_machines)]
    content.add_children(machines)

    jobs = [new_job(i, jobs_pos[i]) for i in range(num_jobs)]
    content.add_children(jobs)

    play_button = PlayButton(pos=(15, 12))
    content.add_child(play_button)

    reset_button = ResetButton(pos=(15 + 35 + 10, 12))
    content.add_child(reset_button)

    chart = SchedulingChart((786, 100), pos=(107, 420))
    content.add_child(chart)

    progress_bar = ProgressBar(800, pos=(100, 410))
    content.add_child(progress_bar)

    time_lbl = Label(font=pygame.font.SysFont("Arial", 24, bold=True), pos=(25, 534))
    content.add_child(time_lbl)

    speed_slider = Slider((150, 25), range=(0, 6), step=0.125, default=1, pos=(835, 15))
    content.add_child(speed_slider)

    speed_lbl = Label(
        f"{speed_slider.value}x",
        font=pygame.font.SysFont("Courier New", 30, bold=True),
        pos=(753, 12),
    )
    content.add_child(speed_lbl)

    def handle_speed_changed(value):
        speed_lbl._text = f"{round(value, 3)}x"
        speed_lbl.repaint()
        width, _ = speed_lbl.get_size()
        speed_lbl.pos = (speed_slider.pos[0] - (width + 10), 12)

    speed_slider.on_value_changed = handle_speed_changed

    # animation variables
    time_stamp = 0

    start_idx = 0
    start_len = len(start_time_list)

    end_idx = 0
    end_len = len(end_time_list)

    def update_time_lbl():
        time_lbl.text = f"{int(time_stamp):02d} / {int(max_time):02d}"

    update_time_lbl()

    def reset():
        nonlocal time_stamp, start_idx, end_idx
        play_button.is_playing = False
        progress_bar.progress = 0.0
        time_stamp = 0
        update_time_lbl()
        start_idx = 0
        end_idx = 0
        for i in range(num_jobs):
            # reset positions
            jobs[i].move_pos(jobs_pos[i])

    reset_button.on_clicked = reset

    try:
        curr_time = time.time()

        # main loop
        while True:
            prev_time = curr_time
            curr_time = time.time()
            delta_time = curr_time - prev_time

            # pygame events
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    return

            # play animation
            if play_button.is_playing:
                speed = speed_slider.value

                if speed != 0:
                    time_stamp += delta_time * speed

                    if end_idx < end_len and time_stamp >= end_time_list[end_idx][0]:
                        _, j, _ = end_time_list[end_idx]
                        jobs[j].move_pos(jobs_pos[j], 0.3 / speed)
                        end_idx += 1

                    if (
                        start_idx < start_len
                        and time_stamp >= start_time_list[start_idx][0]
                    ):
                        _, j, m = start_time_list[start_idx]
                        jobs[j].move_pos(processing_pos[m], 0.3 / speed)
                        start_idx += 1

                    progress_bar.progress = time_stamp / max_time
                    update_time_lbl()

                    if time_stamp >= max_time + 0.3:
                        reset()

            # update components
            content.update()

            # render display
            content.render_update()
            pygame.display.update()

    except KeyboardInterrupt:
        pygame.quit()


if __name__ == "__main__":
    p = read_excel("example.xlsx", "b2:e4", "Book2Problem")
    s = read_excel("example.xlsx", "b7:e9", "Book2Problem")
    visualize(p, s)
