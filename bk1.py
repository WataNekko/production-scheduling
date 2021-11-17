import numpy as np


def read_excel(path: str, cells: str, sheet_name: str = None) -> np.ndarray:
    """Read an Excel file and return an array of cell values."""

    from openpyxl import load_workbook

    # open excel file
    wb = load_workbook(filename=path)
    ws = wb.active if sheet_name is None else wb[sheet_name]

    # read cells
    cells_tuple = ws[cells]
    get_values = np.vectorize(lambda cell: cell.value, otypes=[object])

    return get_values(cells_tuple)


def visualize(process_time, start_time) -> None:
    """
    Visualize the scheduling of a production line problem.

    Parameters:
        process_time: A table representing the processing time of the jobs on different machines,
            where the rows are the jobs and columns are the machines.
        start_time: A table representing the time to start processing the jobs on different machines,
            where the rows are the jobs and columns are the machines.
    """

    import pygame
    import random
    import time

    # input validation
    process_time = np.array(process_time)
    non_none_process_time = process_time[process_time != None]

    start_time = np.array(start_time)
    non_none_start_time = start_time[start_time != None]

    if process_time.shape != start_time.shape or np.any(
        (non_none_process_time < 0) | (non_none_start_time < 0)
    ):
        raise ValueError(
            "The two input tables should be of the same dimensions and all values must be non-negative or None."
        )

    # pygame initialization
    pygame.init()
    pygame.display.set_caption("Scheduling of a production line")

    resolution = (1000, 600)
    screen = pygame.display.set_mode(resolution)

    black = (0, 0, 0)
    white = (255, 255, 255)
    font = pygame.font.SysFont("Courier New", 50, bold=True)

    # components
    class Component:
        def __init__(
            self,
            surface: pygame.Surface,
            is_clipped: bool = True,
            pos: tuple[int, int] = None,
            children: list["Component"] = None,
        ):
            self._surface = surface
            # the is_clipped property allows the parent component to render this child
            # once onto the parent's surface instead of rendering this component every
            # frame onto a different given surface.
            self.is_clipped = is_clipped
            self.pos = pos if pos is not None else (0, 0)
            self._children = children if children is not None else []
            self._parent = None
            for child in self._children:
                child._parent = self
            self._movement = None
            self.repaint()

        def repaint(self):
            for child in self._children:
                # when the parent is repainted, render the children directly onto the parent's
                # surface if they're clipped to avoid repainting them every render frame
                if child.is_clipped:
                    child.render(self._surface)

        def render(self, surface: pygame.Surface, offset: tuple[int, int] = None):
            # offset the position if an offset is given
            pos = (
                (offset[0] + self.pos[0], offset[1] + self.pos[1])
                if offset is not None
                else self.pos
            )
            surface.blit(self._surface, pos)
            for child in self._children:
                # if a child is not clipped, it is rendered seperately onto the given
                # surface with the parent's position as an offset.
                if not child.is_clipped:
                    child.render(surface, self.pos)

        def set_dirty(self):
            pass

        def add_child(self, component: "Component"):
            self._children.append(component)
            # render the new child onto this component's surface if it's clipped
            if component.is_clipped:
                component.render(self._surface)

        def get_size(self):
            return self._surface.get_size()

        def get_pos_center(
            self, target: pygame.Surface | tuple[int, int]
        ) -> tuple[int, int]:
            x, y = self.get_size()
            if isinstance(target, pygame.Surface):
                width, height = target.get_size()
                return round((width - x) / 2), round((height - y) / 2)
            else:
                return round(target[0] - x / 2), round(target[1] - y / 2)

        def update(self):
            if self._movement is not None:
                target, [curr_pos, duration, prev_time] = self._movement

                curr_time = time.time()
                delta_time = curr_time - prev_time
                progress = delta_time / duration

                if progress >= 1:
                    # passed the movement duration
                    self.pos = target
                    self._movement = None
                else:
                    # update pos
                    curr_pos = curr_pos + (target - curr_pos) * progress
                    self.pos = curr_pos.tolist()

                    # update the movement object
                    self._movement[1][0] = curr_pos
                    self._movement[1][1] = duration - delta_time
                    self._movement[1][2] = curr_time

            for child in self._children:
                child.update()

        def move_pos(self, target: tuple[int, int], duration: float = 0.0):
            if duration > 0:
                # record a movement object to be processed in update()
                self._movement = (
                    np.array(target),
                    [np.array(self.pos), duration, time.time()],
                )
            else:
                self.pos = target
                self._movement = None

    class Label(Component):
        def __init__(
            self,
            text: str = "Label",
            color: tuple[int, int, int] = white,
            font: pygame.font.Font = font,
            is_clipped: bool = True,
            pos: tuple[int, int] = None,
            children: list[Component] = None,
        ):
            self._text = text
            self._color = color
            self._font = font
            super().__init__(None, is_clipped=is_clipped, pos=pos, children=children)

        def repaint(self):
            self._surface = self._font.render(self._text, True, self._color)
            super().repaint()

        def set_color(self, color: tuple[int, int, int]):
            self._color = color
            self.repaint()

    class LabeledComponent(Component):
        def __init__(
            self,
            label: str,
            fgcolor: tuple[int, int, int],
            font: pygame.font.Font,
            bgcolor: tuple[int, int, int],
            border_color: tuple[int, int, int],
            size: tuple[int, int],
            is_clipped: bool,
            pos: tuple[int, int],
            children: list[Component],
            is_background_alpha: bool = False,
        ):
            flags = pygame.SRCALPHA if is_background_alpha else 0
            surface = pygame.Surface(size, flags=flags)

            lbl = Label(label, fgcolor, font, is_clipped=False)
            lbl.pos = lbl.get_pos_center(surface)
            self._label = lbl

            self._bgcolor = bgcolor
            self._border_color = border_color

            children = [lbl] + children if children is not None else [lbl]

            super().__init__(surface, is_clipped, pos, children)

    class Machine(LabeledComponent):
        def __init__(
            self,
            label: str = "M",
            fgcolor: tuple[int, int, int] = white,
            font: pygame.font.Font = font,
            bgcolor: tuple[int, int, int] = (128, 25, 59),
            border_color: tuple[int, int, int] = (80, 0, 47),
            size: tuple[int, int] = (124, 124),
            is_clipped: bool = True,
            pos: tuple[int, int] = None,
            children: list[Component] = None,
        ):
            super().__init__(
                label,
                fgcolor,
                font,
                bgcolor,
                border_color,
                size,
                is_clipped,
                pos,
                children,
            )

        def repaint(self):
            self._surface.fill(self._bgcolor)
            if self._border_color is not None:
                pygame.draw.rect(
                    self._surface, self._border_color, self._surface.get_rect(), width=9
                )
            super().repaint()

    class Job(LabeledComponent):
        def __init__(
            self,
            label: str = "J",
            fgcolor: tuple[int, int, int] = white,
            font: pygame.font.Font = font,
            bgcolor: tuple[int, int, int] = (0, 158, 26),
            border_color: tuple[int, int, int] = (0, 75, 12),
            size: tuple[int, int] = (110, 110),
            is_clipped: bool = True,
            pos: tuple[int, int] = None,
            children: list[Component] = None,
        ):
            super().__init__(
                label,
                fgcolor,
                font,
                bgcolor,
                border_color,
                size,
                is_clipped,
                pos,
                children,
                is_background_alpha=True,
            )

        def repaint(self):
            self._surface.fill((0, 0, 0, 0))  # clear the surface
            rect = self._surface.get_rect()
            pygame.draw.ellipse(self._surface, self._bgcolor, rect)
            if self._border_color is not None:
                pygame.draw.ellipse(self._surface, self._border_color, rect, width=6)
            super().repaint()

    class PlayButton:
        pass

    class Slider:
        pass

    class ProgressBar:
        pass

    class SchedulingChart:
        pass

    class RestartButton:
        pass

    # random color
    random.seed(5)

    def get_color():
        return tuple(random.randint(0, 255) for _ in range(3))

    # component list
    num_jobs, num_machines = process_time.shape

    machines = [Machine(f"M{i + 1}") for i in range(num_machines)]
    jobs = [Job(f"J{i + 1}", bgcolor=get_color()) for i in range(num_jobs)]

    components = machines + jobs
    content = Component(pygame.Surface(resolution), children=components)

    try:
        # main loop
        while True:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    pygame.quit()
                    return
                if event.type == pygame.MOUSEBUTTONDOWN:
                    components[0].move_pos(
                        components[0].get_pos_center(pygame.mouse.get_pos()), 0.3
                    )

            # update components
            content.update()
            # for c in components:
            #     c.update()

            # render display
            screen.fill(black)
            content.render(screen)
            # for c in components:
            #     c.render(screen)

            pygame.display.update()

    except KeyboardInterrupt:
        pygame.quit()


if __name__ == "__main__":
    p = read_excel("test.xlsx", "b2:e4")
    s = read_excel("test.xlsx", "b7:e9")
    visualize(p, s)
